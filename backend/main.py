from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request
from starlette.responses import Response
import httpx
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import landscape, letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, PageBreak, HRFlowable)
import io, json, string, math
from datetime import datetime
from typing import List

app = FastAPI(title="FX Financial Report Converter")

class ForceCORSMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        if request.method == "OPTIONS":
            response = Response()
        else:
            try:
                response = await call_next(request)
            except Exception as e:
                response = JSONResponse(status_code=500, content={"detail": str(e)})
        response.headers["Access-Control-Allow-Origin"]  = "*"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = "*"
        return response

app.add_middleware(ForceCORSMiddleware)

SKIP_SHEETS = {"filters","filter","settings","config","metadata","lookup"}


def clean_nan(val):
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return None
    return val

def clean_records(records):
    return [{k: clean_nan(v) for k, v in row.items()} for row in records]


# ── Live FX proxy ─────────────────────────────────────────────────────────────

@app.get("/fx-rate")
async def fx_rate(from_currency: str, to_currency: str):
    try:
        async with httpx.AsyncClient(timeout=8) as client:
            res = await client.get("https://api.frankfurter.dev/v2/rates",
                                   params={"base": from_currency, "quotes": to_currency})
            return res.json()
    except Exception as e:
        raise HTTPException(502, f"Could not fetch live rate: {e}")


# ── Structure detection ───────────────────────────────────────────────────────

def pick_sheet(xl):
    for name in xl.sheet_names:
        if name.strip().lower() not in SKIP_SHEETS:
            return name
    return xl.sheet_names[0]


def detect_structure(contents: bytes):
    xl     = pd.ExcelFile(io.BytesIO(contents))
    sheet  = pick_sheet(xl)
    df_raw = pd.read_excel(io.BytesIO(contents), sheet_name=sheet, header=None)

    header_row_idx = None
    for ri in range(min(8, len(df_raw))):
        vals = [str(v).lower().strip() for v in df_raw.iloc[ri] if pd.notna(v)]
        has_label  = any(k in v for v in vals for k in ["code","account","name","description"])
        has_amount = any(k in v for v in vals for k in ["balance","amount","total","value"])
        if has_label and has_amount:
            header_row_idx = ri
            break

    if header_row_idx is None:
        raise HTTPException(400, f"Could not find column headers in sheet '{sheet}'.")

    df_hdr = pd.read_excel(io.BytesIO(contents), sheet_name=sheet, header=header_row_idx)
    df_hdr.columns = [str(c).strip() for c in df_hdr.columns]

    code_col = next((c for c in df_hdr.columns if "code" in c.lower()), None)
    name_col = next((c for c in df_hdr.columns
                     if any(k in c.lower() for k in ["account name","account","description","name"])
                     and c != code_col), None)
    amt_raw  = [c for c in df_hdr.columns
                if c not in [code_col, name_col]
                and any(k in c.lower() for k in ["balance","amount","total","value"])]

    if not name_col:
        raise HTTPException(400, f"No account name column found. Columns: {list(df_hdr.columns)}")
    if not amt_raw:
        raise HTTPException(400, f"No amount/balance column found. Columns: {list(df_hdr.columns)}")

    n_amt           = len(amt_raw)
    all_col_names   = list(df_hdr.columns)
    amt_col_indices = [all_col_names.index(c) for c in amt_raw]

    branch_names = []
    for ri in range(header_row_idx):
        row_vals   = list(df_raw.iloc[ri])
        candidates = []
        for idx in amt_col_indices:
            if idx < len(row_vals) and pd.notna(row_vals[idx]):
                s = str(row_vals[idx]).strip()
                if s and s.lower() != "nan" and not (s.isdigit() and len(s) == 4):
                    candidates.append(s)
        if len(candidates) == n_amt:
            branch_names = candidates; break
        elif len(candidates) > len(branch_names):
            branch_names = candidates

    if not branch_names:
        branch_names = ["Balance"] if n_amt == 1 else [f"Column {i+1}" for i in range(n_amt)]

    rename_map = dict(zip(amt_raw, branch_names[:n_amt]))
    df = df_hdr.rename(columns=rename_map)

    if code_col: df = df.rename(columns={code_col: "code"})
    else:        df.insert(0, "code", "")
    df = df.rename(columns={name_col: "account_name"})

    keep = ["code","account_name"] + branch_names
    df   = df[[c for c in keep if c in df.columns]].copy()

    df["code"]         = df["code"].fillna("").astype(str).str.strip().str.replace(r"\.0$","",regex=True)
    df["account_name"] = df["account_name"].fillna("").astype(str).str.strip()
    for col in branch_names:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df["row_type"] = df[branch_names].notna().any(axis=1).map({True:"data",False:"header"})
    df = df[~(
        (df["code"] == "") & (df["account_name"] == "") & df[branch_names].isna().all(axis=1)
    )].reset_index(drop=True)

    return df, sheet, branch_names, len(branch_names) > 1


def apply_rate(df, branch_cols, selected_cols, rate):
    df = df.copy()
    for col in branch_cols:
        df[f"{col}__orig"] = df[col]
        df[f"{col}__conv"] = df[col] * rate if col in selected_cols else df[col]
    return df


# ── Single-sheet Excel writer (used inside multi-report workbook) ─────────────

def write_sheet(ws, df, branch_cols, selected_cols, rate, from_cur, to_cur, sheet_name):
    thin = Side(style="thin", color="CCCCCC")
    b    = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_headers = ["Code", "Account Name"]
    col_keys    = []
    for bc in branch_cols:
        col_headers.append(f"{bc}\n({from_cur})")
        if bc in selected_cols:
            col_headers.append(f"{bc}\n({to_cur}) ✓")
        col_keys.append((f"{bc}__orig", f"{bc}__conv" if bc in selected_cols else None))

    n_cols = len(col_headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(1,1).value     = f"{sheet_name}"
    ws.cell(1,1).font      = Font(name="Arial", bold=True, size=14, color="E8D5B7")
    ws.cell(1,1).fill      = PatternFill("solid", start_color="0F3460")
    ws.cell(1,1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(2,1).value = (
        f"Rate: 1 {from_cur} = {rate} {to_cur}   |   "
        f"Converted: {', '.join(selected_cols) or 'none'}   |   "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws.cell(2,1).font      = Font(name="Arial", size=9, color="A8C5DA")
    ws.cell(2,1).fill      = PatternFill("solid", start_color="16213E")
    ws.cell(2,1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.append([])

    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(4, ci, h)
        c.font      = Font(name="Arial", bold=True, size=10, color="E8D5B7")
        c.fill      = PatternFill("solid", start_color="1A1A2E")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = b
    ws.row_dimensions[4].height = 32

    data_count = 0
    for _, row in df.iterrows():
        r      = ws.max_row + 1
        is_hdr = row["row_type"] == "header"
        if not is_hdr: data_count += 1
        row_fill = (PatternFill("solid", start_color="2D2B55") if is_hdr else
                    PatternFill("solid", start_color="F5F3FF") if data_count % 2 == 0 else None)

        c1 = ws.cell(r, 1, "" if is_hdr else str(row["code"]))
        c1.font = Font(name="Arial", size=9, color="C4B5FD" if is_hdr else "9CA3AF")
        c1.border = b; c1.alignment = Alignment(vertical="center")
        if row_fill: c1.fill = row_fill

        c2 = ws.cell(r, 2, str(row["account_name"]))
        c2.font = Font(name="Arial", bold=is_hdr, size=10, color="C4B5FD" if is_hdr else "1F2937")
        c2.border = b
        c2.alignment = Alignment(vertical="center", indent=0 if is_hdr else 1)
        if row_fill: c2.fill = row_fill

        ci = 3
        for orig_key, conv_key in col_keys:
            v   = row.get(orig_key)
            val = float(v) if pd.notna(v) else None
            c   = ws.cell(r, ci, val)
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = b
            c.font = Font(name="Arial", bold=is_hdr, size=9, color="C4B5FD" if is_hdr else "374151")
            if row_fill: c.fill = row_fill
            ci += 1
            if conv_key:
                cv   = row.get(conv_key)
                cval = float(cv) if pd.notna(cv) else None
                cc   = ws.cell(r, ci, cval)
                cc.number_format = "#,##0.00"
                cc.alignment = Alignment(horizontal="right", vertical="center")
                cc.border = b
                cc.font = Font(name="Arial", bold=is_hdr, size=9, color="86EFAC" if is_hdr else "065F46")
                cc.fill = row_fill if row_fill else PatternFill("solid", start_color="F0FDF4")
                ci += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 42
    alpha = list(string.ascii_uppercase)
    for i in range(n_cols - 2):
        ws.column_dimensions[alpha[2 + i]].width = 17

    return n_cols


# ── Report type detection ─────────────────────────────────────────────────────

# Maps keywords found in filename or sheet name → report type label + tab suffix
REPORT_TYPE_MAP = [
    (["profit","loss","p&l","pnl","income","revenue"],  "P&L",           "P&L"),
    (["balance","sheet","bs"],                           "Balance Sheet", "Balance Sheet"),
    (["aged","ar","receivable"],                         "Aged AR",       "Aged AR"),
    (["aged","ap","payable"],                            "Aged AP",       "Aged AP"),
    (["trial","balance","tb"],                           "Trial Balance", "Trial Balance"),
    (["cash","flow"],                                    "Cash Flow",     "Cash Flow"),
]

def infer_report_type(filename: str, sheet_name: str) -> tuple[str, str]:
    """
    Returns (report_type_label, tab_suffix) by scanning filename and sheet name.
    Falls back to the filename if nothing matches.
    """
    haystack = f"{filename} {sheet_name}".lower()
    for keywords, label, suffix in REPORT_TYPE_MAP:
        # Require at least 2 keywords to match for multi-keyword types (aged ar/ap, cash flow)
        # For single-distinguishing types just one match is enough
        matches = sum(1 for kw in keywords if kw in haystack)
        if matches >= 1:
            return label, suffix
    return "Report", filename  # fallback


def make_tab_name(report_type_suffix: str, to_cur: str, existing_tabs: list[str]) -> str:
    """
    Builds a clean Excel tab name like 'P&L USD', 'Balance Sheet USD'.
    Ensures uniqueness and stays within Excel's 31-char limit.
    """
    base = f"{report_type_suffix} {to_cur}"[:29].strip()
    name = base
    counter = 2
    while name in existing_tabs:
        name = f"{base} ({counter})"
        counter += 1
    return name


# ── Summary sheet ─────────────────────────────────────────────────────────────

def write_summary_sheet(ws, reports_meta, rate, from_cur, to_cur):
    """
    Writes a clean summary sheet. One row per report — no Grand Total.
    Columns: Report Name | Report Type | Local Currency | Local Total | to_cur Total | Rate Used

    reports_meta: list of dicts with keys:
        filename, sheet_name, report_type, total_orig, total_conv, rate
    """
    thin = Side(style="thin", color="CCCCCC")
    b    = Border(left=thin, right=thin, top=thin, bottom=thin)
    inv  = round(1 / rate, 4) if rate else 0

    # ── Row 1: Title ──
    ws.merge_cells("A1:F1")
    ws.cell(1,1).value     = "Financial Reports — Consolidated Summary"
    ws.cell(1,1).font      = Font(name="Arial", bold=True, size=16, color="E8D5B7")
    ws.cell(1,1).fill      = PatternFill("solid", start_color="0F3460")
    ws.cell(1,1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Row 2: Rate info ──
    ws.merge_cells("A2:F2")
    ws.cell(2,1).value = (
        f"1 {from_cur} = {rate:.4f} {to_cur}   |   "
        f"1 {to_cur} = {inv:.4f} {from_cur}   |   "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   "
        f"{len(reports_meta)} report(s)"
    )
    ws.cell(2,1).font      = Font(name="Arial", size=10, color="A8C5DA")
    ws.cell(2,1).fill      = PatternFill("solid", start_color="16213E")
    ws.cell(2,1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.append([])  # blank spacer row 3

    # ── Row 4: Column headers ──
    headers = [
        "Report Name",
        "Report Type",
        "Local Currency",
        f"Local Total ({from_cur})",
        f"{to_cur} Total",
        "Rate Used",
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(4, ci, h)
        c.font      = Font(name="Arial", bold=True, size=11, color="E8D5B7")
        c.fill      = PatternFill("solid", start_color="1A1A2E")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = b
    ws.row_dimensions[4].height = 24

    # Freeze header rows so they stay visible when scrolling
    ws.freeze_panes = "A5"

    # ── Data rows — one per report, NO grand total ──
    for i, meta in enumerate(reports_meta):
        r        = 5 + i
        alt_fill = PatternFill("solid", start_color="F0F4FF") if i % 2 == 0 else None

        row_data = [
            (1, meta["filename"],      False),   # Report Name
            (2, meta["report_type"],   False),   # Report Type
            (3, from_cur,              False),   # Local Currency
            (4, meta["total_orig"],    True),    # Local Total
            (5, meta["total_conv"],    True),    # to_cur Total
            (6, meta["rate"],          False),   # Rate Used
        ]

        for ci, val, is_num in row_data:
            c = ws.cell(r, ci, val)
            c.border    = b
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(
                vertical="center",
                horizontal="right" if is_num or ci == 6 else "left"
            )
            if is_num:
                c.number_format = "#,##0.00"
            if ci == 6:
                c.number_format = "0.00000"  # show rate to 5 decimal places
            if alt_fill:
                c.fill = alt_fill

        ws.row_dimensions[r].height = 20

    # ── Auto-size columns ──
    col_widths = [35, 20, 18, 22, 22, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Multi-report Excel ────────────────────────────────────────────────────────

def build_multi_excel(reports, rate, from_cur, to_cur):
    """
    Builds one workbook:
      - Tab 1: Summary (one row per report, no grand total, rate header)
      - Tab N: One tab per report, named '{Report Type} {to_cur}'
                e.g. 'P&L USD', 'Balance Sheet USD'

    Conversion is always: converted = local_amount * rate
    """
    wb = Workbook()

    # Enrich each report with inferred type and clean tab name
    existing_tabs = ["Summary"]
    for rep in reports:
        report_type, suffix = infer_report_type(rep["filename"], rep["sheet_name"])
        rep["report_type"] = report_type
        rep["tab_name"]    = make_tab_name(suffix, to_cur, existing_tabs)
        rep["rate"]        = rate
        existing_tabs.append(rep["tab_name"])

    # Summary sheet first
    ws_summary        = wb.active
    ws_summary.title  = "Summary"
    write_summary_sheet(ws_summary, reports, rate, from_cur, to_cur)

    # One sheet per report with smart tab name
    for rep in reports:
        ws = wb.create_sheet(title=rep["tab_name"])
        write_sheet(ws, rep["df"], rep["branch_cols"], rep["selected_cols"],
                    rate, from_cur, to_cur, rep["filename"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Multi-report PDF ──────────────────────────────────────────────────────────

def build_multi_pdf(reports, rate, from_cur, to_cur):
    buf  = io.BytesIO()
    # Use landscape if any report has >2 branch cols
    has_wide = any(len(r["branch_cols"]) > 2 for r in reports)
    page = landscape(letter) if has_wide else letter

    doc = SimpleDocTemplate(buf, pagesize=page,
                             leftMargin=0.5*inch, rightMargin=0.5*inch,
                             topMargin=0.6*inch,  bottomMargin=0.6*inch)
    styles = getSampleStyleSheet()

    cover_title = ParagraphStyle("CT", parent=styles["Title"], fontSize=22,
                                  textColor=colors.HexColor("#0F3460"), spaceAfter=6,
                                  fontName="Helvetica-Bold", alignment=1)
    cover_sub   = ParagraphStyle("CS", parent=styles["Normal"], fontSize=11,
                                  textColor=colors.HexColor("#374151"), spaceAfter=4,
                                  fontName="Helvetica", alignment=1)
    section_ttl = ParagraphStyle("ST", parent=styles["Heading1"], fontSize=13,
                                  textColor=colors.HexColor("#0F3460"), spaceAfter=4,
                                  spaceBefore=6, fontName="Helvetica-Bold")
    tbl_sub     = ParagraphStyle("TS", parent=styles["Normal"], fontSize=8,
                                  textColor=colors.HexColor("#374151"), spaceAfter=8)

    story = []

    # ── Cover page ──
    story.append(Spacer(1, 1.2*inch))
    story.append(Paragraph("Financial Reports Package", cover_title))
    story.append(Paragraph("Currency Conversion Summary", cover_sub))
    story.append(Spacer(1, 0.3*inch))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#0F3460")))
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph(
        f"Exchange Rate: &nbsp; 1 {from_cur} = {rate} {to_cur}", cover_sub))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}", cover_sub))
    story.append(Paragraph(f"Total reports: {len(reports)}", cover_sub))
    story.append(Spacer(1, 0.4*inch))

    # Summary table on cover
    sum_data = [["Report", f"Original ({from_cur})", f"Converted ({to_cur})", "Difference"]]
    grand_o, grand_c = 0, 0
    for rep in reports:
        o = rep["total_orig"] or 0
        c = rep["total_conv"] or 0
        sum_data.append([rep["filename"], f"{o:,.2f}", f"{c:,.2f}", f"{c-o:,.2f}"])
        grand_o += o; grand_c += c
    sum_data.append(["GRAND TOTAL", f"{grand_o:,.2f}", f"{grand_c:,.2f}", f"{grand_c-grand_o:,.2f}"])

    sw = [2.8*inch, 1.4*inch, 1.4*inch, 1.2*inch]
    stbl = Table(sum_data, colWidths=sw)
    stbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),  (-1,0),  colors.HexColor("#1A1A2E")),
        ("TEXTCOLOR",     (0,0),  (-1,0),  colors.HexColor("#E8D5B7")),
        ("FONTNAME",      (0,0),  (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),  (-1,-1), 9),
        ("ALIGN",         (1,0),  (-1,-1), "RIGHT"),
        ("ALIGN",         (0,0),  (0,-1),  "LEFT"),
        ("VALIGN",        (0,0),  (-1,-1), "MIDDLE"),
        ("FONTNAME",      (0,1),  (-1,-2), "Helvetica"),
        ("ROWBACKGROUND", (0,1),  (-1,-2), [colors.white, colors.HexColor("#F0F4FF")]),
        ("BACKGROUND",    (0,-1), (-1,-1), colors.HexColor("#0F3460")),
        ("TEXTCOLOR",     (0,-1), (-1,-1), colors.HexColor("#E8D5B7")),
        ("FONTNAME",      (0,-1), (-1,-1), "Helvetica-Bold"),
        ("GRID",          (0,0),  (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",    (0,0),  (-1,-1), 5),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 5),
    ]))
    story.append(stbl)
    story.append(PageBreak())

    # ── One section per report ──
    for rep in reports:
        story.append(Paragraph(rep["filename"], section_ttl))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#C4B5FD")))
        story.append(Paragraph(
            f"Rate: 1 {from_cur} = {rate} {to_cur} &nbsp;|&nbsp; "
            f"Converted columns: {', '.join(rep['selected_cols']) or 'none'}",
            tbl_sub))

        df         = rep["df"]
        branch_cols= rep["branch_cols"]
        sel        = rep["selected_cols"]

        avail_w = page[0] - inch
        name_w  = 2.0*inch; code_w = 0.65*inch
        n_amt   = len(branch_cols) + len(sel)
        amt_w   = max(0.6*inch, min(1.05*inch, (avail_w - name_w - code_w) / max(n_amt, 1)))

        col_headers = ["Code", "Account Name"]
        col_widths  = [code_w, name_w]
        for bc in branch_cols:
            col_headers.append(f"{bc[:16]}\n({from_cur})")
            col_widths.append(amt_w)
            if bc in sel:
                col_headers.append(f"{bc[:16]}\n({to_cur})✓")
                col_widths.append(amt_w)

        rows = [col_headers]
        for _, row in df.iterrows():
            is_hdr = row["row_type"] == "header"
            tr = ["" if is_hdr else str(row["code"]), str(row["account_name"])]
            for bc in branch_cols:
                v = row.get(f"{bc}__orig")
                tr.append(f"{float(v):,.2f}" if pd.notna(v) else "")
                if bc in sel:
                    cv = row.get(f"{bc}__conv")
                    tr.append(f"{float(cv):,.2f}" if pd.notna(cv) else "")
            rows.append(tr)

        tbl = Table(rows, colWidths=col_widths, repeatRows=1)
        cmd = [
            ("BACKGROUND",    (0,0),  (-1,0),  colors.HexColor("#1A1A2E")),
            ("TEXTCOLOR",     (0,0),  (-1,0),  colors.HexColor("#E8D5B7")),
            ("FONTNAME",      (0,0),  (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),  (-1,-1), 7),
            ("ALIGN",         (0,0),  (-1,0),  "CENTER"),
            ("VALIGN",        (0,0),  (-1,-1), "MIDDLE"),
            ("FONTNAME",      (0,1),  (-1,-1), "Helvetica"),
            ("ALIGN",         (2,1),  (-1,-1), "RIGHT"),
            ("GRID",          (0,0),  (-1,-1), 0.3, colors.HexColor("#E5E7EB")),
            ("TOPPADDING",    (0,0),  (-1,-1), 3),
            ("BOTTOMPADDING", (0,0),  (-1,-1), 3),
            ("ROWBACKGROUND", (0,1),  (-1,-1), [colors.white, colors.HexColor("#F5F3FF")]),
        ]
        for ri, row in enumerate(df.itertuples(), start=1):
            if row.row_type == "header":
                cmd += [
                    ("BACKGROUND", (0,ri), (-1,ri), colors.HexColor("#2D2B55")),
                    ("TEXTCOLOR",  (0,ri), (-1,ri), colors.HexColor("#C4B5FD")),
                    ("FONTNAME",   (0,ri), (-1,ri), "Helvetica-Bold"),
                ]
        tbl.setStyle(TableStyle(cmd))
        story.append(tbl)
        story.append(PageBreak())

    doc.build(story)
    return buf.getvalue()


# ── Request helpers ───────────────────────────────────────────────────────────

async def parse_one(file: UploadFile, exchange_rate: float, selected_cols_json: str):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, f"{file.filename}: only .xlsx / .xls supported.")
    contents = await file.read()
    df, sheet, branch_cols, is_multi = detect_structure(contents)
    sel = json.loads(selected_cols_json) if selected_cols_json else branch_cols
    sel = [c for c in sel if c in branch_cols] or branch_cols
    df  = apply_rate(df, branch_cols, sel, exchange_rate)
    data_rows = df[df["row_type"] == "data"]
    total_orig = clean_nan(float(data_rows[[f"{bc}__orig" for bc in sel]].sum(skipna=True).sum()))
    total_conv = clean_nan(float(data_rows[[f"{bc}__conv" for bc in sel]].sum(skipna=True).sum()))
    return {
        "df": df, "sheet_name": sheet, "branch_cols": branch_cols,
        "selected_cols": sel,
        "filename": file.filename.replace(".xlsx","").replace(".xls",""),
        "total_orig": total_orig, "total_conv": total_conv,
        "rate": exchange_rate,
    }


# ── Routes ────────────────────────────────────────────────────────────────────

@app.post("/detect")
async def detect(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx / .xls files supported.")
    contents = await file.read()
    df, sheet, branch_cols, is_multi = detect_structure(contents)
    return {"sheet_name": sheet, "branch_cols": branch_cols,
            "is_multi": is_multi, "row_count": len(df[df["row_type"]=="data"])}


@app.post("/preview")
async def preview(
    file:          UploadFile = File(...),
    exchange_rate: float      = Form(...),
    from_currency: str        = Form("USD"),
    to_currency:   str        = Form("EUR"),
    selected_cols: str        = Form("[]"),
):
    rep = await parse_one(file, exchange_rate, selected_cols)
    df  = rep["df"]
    data_rows  = df[df["row_type"] == "data"]
    branch_cols= rep["branch_cols"]
    sel        = rep["selected_cols"]

    totals = {}
    for bc in branch_cols:
        totals[f"{bc}__orig"] = clean_nan(float(data_rows[f"{bc}__orig"].sum(skipna=True)))
        if bc in sel:
            totals[f"{bc}__conv"] = clean_nan(float(data_rows[f"{bc}__conv"].sum(skipna=True)))

    preview_keys = (["code","account_name","row_type"] +
                    [k for bc in branch_cols
                     for k in ([f"{bc}__orig",f"{bc}__conv"] if bc in sel else [f"{bc}__orig"])])
    return {
        "sheet_name": rep["sheet_name"], "branch_cols": branch_cols,
        "selected_cols": sel, "is_multi": len(branch_cols)>1,
        "rows": len(data_rows), "totals": totals,
        "from_currency": from_currency, "to_currency": to_currency,
        "exchange_rate": exchange_rate,
        "preview": clean_records(df[preview_keys].head(15).to_dict("records")),
    }


@app.post("/download/xlsx")
async def download_xlsx(
    files:         List[UploadFile] = File(...),
    exchange_rate: float            = Form(...),
    from_currency: str              = Form("USD"),
    to_currency:   str              = Form("EUR"),
    selected_cols: str              = Form("[]"),
):
    reports = [await parse_one(f, exchange_rate, selected_cols) for f in files]
    data    = build_multi_excel(reports, exchange_rate, from_currency, to_currency)
    fname   = "financial_report.xlsx" if len(reports)==1 else "financial_reports_package.xlsx"
    return StreamingResponse(io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.post("/download/pdf")
async def download_pdf(
    files:         List[UploadFile] = File(...),
    exchange_rate: float            = Form(...),
    from_currency: str              = Form("USD"),
    to_currency:   str              = Form("EUR"),
    selected_cols: str              = Form("[]"),
):
    reports = [await parse_one(f, exchange_rate, selected_cols) for f in files]
    data    = build_multi_pdf(reports, exchange_rate, from_currency, to_currency)
    fname   = "financial_report.pdf" if len(reports)==1 else "financial_reports_package.pdf"
    return StreamingResponse(io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.get("/health")
async def health():
    return {"status": "ok"}
